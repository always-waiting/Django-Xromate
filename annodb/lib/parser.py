# encoding: utf-8
import os
import re
import StringIO
from ftplib import FTP
from lxml import etree
from Queue import Queue
from threading import Thread
import json
import urllib3
from openpyxl import load_workbook
urllib3.disable_warnings()

def test():
    print "测试导入"

# 把所有解析文件的处理都写到这里

class ParseMim2gene(object):
    """
    解析mim2gene.txt文件一条记录，一条记录为一行
    """
    def __init__(self, chunk):
        self.record = {}
        self.parse_mim2gene(chunk)

    def parse_mim2gene(self, chunk):
        items = chunk.split("\t")
        self.mimNumber = int(items[0])
        self.type = items[1]
        try:
            self.approvedGeneSymbol = items[3]
        except:
            self.approvedGeneSymbol = None

class ParseRefFlat(object):
    """
    解析整个refFlat.txt文件
    """
    def __init__(self, fp, debug=False):
        if not os.path.exists(fp): raise Exception("%s not exists" % fp)
        self.gene = {}
        self.debug = debug
        self.filename = fp
        self.parse_refflat()

    def parse_refflat(self):
        if self.debug: print "Parsing filename %s" % self.filename
        with open(self.filename) as f:
            for line in f:
                match = re.search("(\S+)\t(\S+)\tchr(\S+)\t(\S)\t(\d+)\t(\d+)\t(\d+)\t(\d+)\t(\d+)\t(\S+)\t(\S+)", line)
                #match = re.search("(\S+)\t\S+\tchr(\S+)\t(\S)\t(\d+)\t(\d+)", line)
                if match:
                    pos = {
                        'geneName': match.group(1), 'name': match.group(2), 'chrom': match.group(3), 'strand': match.group(4),
                        'txStart': int(match.group(5)), 'txEnd': int(match.group(6)), 'cdsStart': int(match.group(7)),
                        'cdsEnd': int(match.group(8)), 'exonCount': int(match.group(9)),
                        'exonStarts': list(map(int, match.group(10).rstrip(',').split(','))),
                        'exonEnds': list(map(int, match.group(11).rstrip(',').split(',')))
                        #'name': match.group(1), 'chr': match.group(2), 'strand': match.group(3),
                        #'start': int(match.group(4)), 'end': int(match.group(5))
                    }
                    if re.match("^([1-9]|1[0-9]|2[0-2]|[XY])$", pos['chrom']):
                        if self.gene.has_key(pos['geneName']):
                            self.gene[pos['geneName']].append(pos)
                        else:
                            self.gene[pos['geneName']] = [pos]
        if self.debug: print "Parsing filename Done"
    def make_changes(self, symbol):
        if self.debug: print "Make changes for %s(TO OMIM Genemaps)" % symbol
        symbolist = list(map(str,range(1,23)))
        symbolist.extend(['X','Y'])
        symbol2chr = dict(zip(symbolist, list(range(1,25))))
        try:
            gene = self.gene[symbol]
        except KeyError,e:
            print "Not found key %s" % e
            return {}
        chrom = gene[0]['chrom']; strand = gene[0]['strand']
        start = min(list(map(lambda x: x['txStart'], gene)))
        end = max(list(map(lambda x: x['txEnd'], gene)))
        changes = { # 这里起始位置用所有坐标最小值，终止位置为最大值,力求包含所有区域
            "approvedGeneSymbol": symbol,
            'chromosome': symbol2chr[chrom],
            'chromosomeSymbol': chrom,
            'chromosomeLocationStart': start,
            'chromosomeLocationEnd': end,
            'chromosomeLocationStrand': strand
        }
        return changes



class ParseGenemap(object):
    """
    解析genemap.txt文件的一条记录，一条记录即为一行
    """
    def __init__(self, chunk):
        self.record = {}
        self.parse_genemap(chunk)

    def parse_genemap(self, chunk):
        item = chunk.split("\t")
        (chrnum, seqid) = item[0].split(".")
        chr_symbols = list(range(23))
        chr_symbols.extend(["X","Y"])
        self.record['chromosome'] = int(chrnum)
        self.record['chromosomeSymbol'] = str(chr_symbols[int(chrnum)])
        self.record['sequenceID'] = int(seqid)
        headers = 'month day year cytoLocation geneSymbols confidence \
            title mimNumber mappingMethod comments disorders \
            mouseGeneSymbol references'.split()
        for i in range(len(headers)):
            try:
                if headers[i] == "mimNumber":
                    self.record[headers[i]] = int(item[i+1])
                elif headers[i] == 'confidence':
                    self.record[headers[i]] = item[i+1].upper()
                elif re.search("month|day|year",headers[i]):
                    continue
                else:
                    self.record[headers[i]] = item[i+1]
            except IndexError:
                pass
            except:
                raise Exception("在解析genemap.txt时发生未知错误")

class ParseMorbidmap(object):
    """
    解析morbidmap.txt文件的一条记录，一条记录为一行
    """
    def __init__(self, chunk):
        self.record = {}
        self.parse_morbidmap(chunk)

    def parse_morbidmap(self, chunk):
        items = chunk.split("\t")
        try:
            phenotype_desc = items[0]
            self.record['geneSymbols'] = items[1]
            self.record['mimNumber'] = int(items[2])
            self.record['cytoLocation'] = items[3]
            if re.match("(.*)\,\s(\d{6})\s?\((\d+)\)", phenotype_desc):
                get = re.match("(.*)\,\s(\d{6})\s?\((\d+)\)", phenotype_desc)
                self.record['phenotype'] = get.group(1)
                self.record['phenotypeMimNumber'] = int(get.group(2))
                self.record['phenotypeMappingKey'] = int(get.group(3))
            elif re.match("(.*)\,\s(\d{6})", phenotype_desc):
                get = re.match("(.*)\,\s(\d{6})", phenotype_desc)
                self.record['phenotype'] = get.group(1)
                self.record['phenotypeMimNumber'] = int(get.group(2))
            elif re.match("(.*)\s?\((\d+)\)", phenotype_desc):
                get = re.match("(.*)\s?\((\d+)\)", phenotype_desc)
                self.record['phenotype'] = get.group(1)
                self.record['phenotypeMimNumber'] = self.record['mimNumber']
                self.record['phenotypeMappingKey'] = int(get.group(2))
            else:
                self.record['phenotype'] = phenotype_desc
                self.record['phenotypeMimNumber'] = self.record['mimNumber']
        except Exception, e:
            print e
            print "出现错误"
            print chunk


class ParseEntryDownload(object):
    """
    输入文件第一列为需要下载的mimNumber
    """
    def __init__(self, filepath):
        self.mims = []
        if not os.path.exists(filepath): raise Exception(filepath + " not exists")
        self.parse_entry_downloadlist(filepath)

    def parse_entry_downloadlist(self, filepath):
        with open(filepath) as f:
            for line in f:
                items = line.rstrip().lstrip().split()
                if re.match("\d{6}", items[0]):
                    self.mims.append(int(items[0]))



class ParseEntry(object):
    """
    解析omim.txt文件的一条记录以**RECORD**作为分割
    """
    def __init__(self, chunk):
        self.record = {}
        self.parse_entry(chunk)

    def parse_entry(self, chunk):
        for field in chunk.split("*FIELD*"):
            if not field: continue
            if field == "\n": continue
            content = StringIO.StringIO(field)
            header = content.next().rstrip().lstrip()
            if header.upper().find("NO") != -1:
                self.header_no(content);
            elif header.upper().find("TI") != -1:
                self.header_ti(content)
            elif header.upper().find("TX") != -1:
                self.header_tx(content)
            elif header.upper().find("SA") != -1:
                self.header_sa(content)
            elif header.upper().find("RF") != -1:
                self.header_rf(content)
            elif header.upper().find("CS") != -1:
                self.header_cs(content)
            elif header.upper().find("CN") != -1:
                self.header_cn(content)
            elif header.upper().find("CD") != -1:
                self.header_cd(content)
            elif header.upper().find("ED") != -1:
                self.header_ed(content)
            elif header.upper().find("AV") != -1:
                self.header_av(content)
            else:
                raise Exception(header + u"没有对应的解析器!!")

    def header_cs(self, content):
        #pass
        record_cs = {}
        record_cs['mimNumber'] = self.record['mimNumber']
        record_cs['prefix'] = self.record['prefix']
        record_cs['preferredTitle'] = self.record['title']['preferredTitle']
        record_cs['oldFormatExists'] = False
        text = content.read().lstrip().rstrip()
        if not text: return
        items = [ re.sub("\s+", " ", n.replace(":","")).lstrip().rstrip() for n in re.split("([A-z, ]+\:\n)",text)]
        if not items[0]: items.pop(0)
        number = len(items)/2.
        if number.is_integer():
            for i in range(int(number)):
                key = items[2*i]
                cs_content = items[2*i+1]
                if re.match(\
                    "INHERITANCE|GROWTH|HEAD AND NECK|CARDIOVASCULAR|RESPIRATORY|CHEST|ABDOMEN|GENITOURINARY|SKELETAL|SKIN\, NAILS\, HAIR|\
                    MUSCLE.*SOFT.*TISSUE|NEUROLOGIC|VOICE|METABOLIC FEATURES|ENDOCRINE FEATURES|HEMATOLOGY|IMMUNOLOGY|NEOPLASIA|PRENATAL MANIFESTATIONS|\
                    LABORATORY ABNORMALITIES|MISCELLANEOUS|MOLECULAR BASIS",\
                    key):
                    if not cs_content: continue
                    key_attr = key.title().replace(" ","")
                    key_attr = key_attr[0].lower() + key_attr[1:]
                    innercontent = re.split('(\[[^;]+\]\;)', cs_content)
                    if not innercontent[0]: innercontent.pop(0)
                    if len(innercontent) % 2: #有键的介绍
                        key_desc = innercontent.pop(0)
                    else:
                        key_desc = ""
                    record_cs[key_attr+'Exists'] = True
                    if key_desc: record_cs[key_attr] = key_desc
                    if not len(innercontent): continue
                    for i in range(int(len(innercontent)/2)):
                        #sub_key = innercontent[2*i]
                        #sub_conntent = innercontent[2*i+1]
                        sub_attr = key_attr + re.sub("\[|\];|,","",innercontent[2*i].title()).replace(" ","")
                        record_cs[sub_attr+"Exists"] = True
                        record_cs[sub_attr] = innercontent[2*i+1].lstrip().rstrip()
                else:
                    record_cs['oldFormatExists'] = True
                    try:
                        record_cs['oldFormat'][key] = cs_content
                    except:
                        record_cs['oldFormat'] = {key:cs_content}
            self.record['clinicalSynopsis'] = record_cs
            self.record['clinicalSynopsisExists'] = True
        else:
            #raise Exception(u"CS域的值非偶数")
            print u"CS域的值非偶数,因此不导入数据库".encode('utf-8').strip()



    def header_av(self, content):
        self.record['allelicVariantExists'] = True
        rawList = re.split("\n\.(\d{4})\n","\n" + content.read().lstrip().rstrip())
        if not rawList[0]: rawList.pop(0)
        number = int(len(rawList)/2)
        allelicVariantList = []
        for i in range(number):
            allele = rawList[2*i]
            text = rawList[2*i+1].lstrip().rstrip()
            hashlist = {'number': int(allele)}
            try:
                index = text.index("\n")
                name = text[0:index]
                splittext = re.split("\n\n",text[index+1:])
                allelenamestr = splittext.pop(0)
                allelenamelist = re.split(";;", allelenamestr)
                if len(allelenamelist) > 1:
                    hashlist['alternativeNames'] = ";;".join(allelenamelist[:-1])
                    hashlist['mutations'] = allelenamelist[-1]
                else:
                    hashlist['mutations'] = allelenamelist[-1]
                hashlist['text'] = "\n".join([ n.lstrip().rstrip().replace("\n"," ") for n in splittext])
            except:#只有一行，应该是moved或者removed
                name = text
            if name.lower().find("removed") != -1:
                status = 'removed'
            elif name.lower().find("moved") != -1:
                status = "moved"
            else:
                status = 'live'
            hashlist['name'] = name
            hashlist['status'] = status
            #print hashlist
            allelicVariantList.append(hashlist)
            self.record['allelicVariantList'] = allelicVariantList

    def header_ed(self, content):
        self.record['editHistory'] = content.read().lstrip().rstrip()

    def header_cd(self,content):
        self.record['createDate'] = content.read().lstrip().rstrip()

    def header_cn(self,content):
        self.record['contributors'] = content.read().lstrip().rstrip()

    def header_rf(self,content):
        rfrecords = []
        rflist = [n.replace("\n"," ") for n in re.split("\r?\n\r?\n", content.read())]
        for rf in rflist:
            if not rf:
                continue
            #get = re.match("(\d+)\.\s+([^:]+)\:\s+([^\.]+)[\.\?]+[\s']+(.*)", rf)
            # 不再添加source
            get = re.match("(\d+)\.\s+([^:]+)\:\s+(.+)", rf)
            try:
                refnumber = int(get.group(1))
                authors = get.group(2)
                title = get.group(3)
                source = ""
            except:
                print "$$$$ ->", rf
                #get = re.match("(\d+)\.\s+([^:]+)\:\s+([^\.]+(\.[^\.]+)*)[\.\?]+[\s']+(\D.*)",rf)
                continue
            rfrecords.append({
                'reference' : {
                    'referenceNumber' : refnumber,
                    'authors'         : authors,
                    'title'           : title,
                    'source'          : source
                }
            })
            #if int(get.group(1)) == 1: print rf
        self.record['referenceList'] = rfrecords

    def header_sa(self,content):
        self.record['seeAlso'] = content.read().lstrip().rstrip().replace("\n"," ")

    def header_tx(self,content):
        self.record['textSectionList'] = []
        all = content.read()
        values = [ n.lstrip().rstrip().replace("\n", " ") for n in re.split("\n[A-Z\s]+\n", all)]
        if not values[0]: values.pop(0)
        keys = [ n.lstrip().rstrip().replace("\n", "") for n in re.findall("\n[A-Z\s]+\n", all)]
        if not len(keys):
            self.record['textSectionList'].append({
                'textSectionName'    : 'text',
                'textSectionTitle'   : 'Text',
                'textSectionContent' : values[0]
            })
            return
        for i in range(len(keys)):
            textsectionname = keys[i].title().replace(" ","")
            textsectionname = textsectionname[0].lower() + textsectionname[1:]
            self.record['textSectionList'].append({
                'textSectionName'    : textsectionname,
                'textSectionTitle'   : keys[i].title(),
                'textSectionContent' : values[i]
            })

    def header_ti(self,content):
        th = content.next().rstrip().lstrip()
        fieldcontent = {}
        get = re.match("(\S?)\d+\s(.*)",th)
        self.record['prefix'] = get.group(1)
        self.record['title'] = { 'preferredTitle': get.group(2) }
        move = re.match("MOVED TO (\d+)", get.group(2))
        remove = re.match("REMOVED", get.group(2))
        if remove:
            self.record['status'] = "removed"
        elif move:
            self.record['status'] = "moved"
            self.record['movedTo'] = int(move.group(1))
        else:
            self.record['status'] = 'live'
        remind = content.read().replace("\n"," ")
        if remind: self.record['title']['alternativeTitles'] = remind

    def header_no(self,content):
        mimnumber = content.next().rstrip().lstrip()
        self.record['mimNumber'] = int(mimnumber)
        print "Go throught mim", self.record['mimNumber']

class ParseClinVar(object):
    """
    解析ClinVarFullRelease_00-latest.xml文件，生成可以导入mongo数据库的dict格式。
    这个一个大型文件，需要用到一定的处理方式来确保运行效率
    """
    def __init__(self, filename, debug=False):
        if not os.path.exists(filename): raise Exception("%s not exist!" % filename)
        self.filename = filename
        self.debug = debug
        self._count = 0
    def __enter__(self):
        return self;
    def __exit__(self ,type, value, traceback):
        return False
    def __iter__(self):
        try:
            if self.fileiter.closed:
                self.fileiter = open(self.filename)
            else:
                self.fileiter.close()
                self.fileiter = open(self.filename)
        except AttributeError,e:
            self.fileiter = open(self.filename)
        return self._parseclinvar()

    def _parseclinvar(self):
        parseflag = False
        buf = ""
        for line in self.fileiter:
            if re.match("<ClinVarSet .*>", line):
                parseflag = True
                buf += line
            elif re.match("</ClinVarSet>", line):
                parseflag = False
                buf+= line
                # 开始解析buf生成结果，然后yield回去
                xmltree = etree.fromstring(buf)
                yield self._generate(xmltree)
                # 初始化
                buf = ""
            elif parseflag:
                buf += line
            else:
                pass
                #print line
    def _generate(self, clinvar):
        """
        传递一条记录的xml tree对象，解析后生成需要的字典
        """
        self._count += 1
        if not self._count%2000: print "parse count: %s" % self._count
        RCVA = clinvar.find("ReferenceClinVarAssertion")
        CVAs = clinvar.findall("ClinVarAssertion")
        CVA = RCVA.find("ClinVarAccession")
        rcv_accession = CVA.attrib['Acc']
        Measure = RCVA.find('MeasureSet').find('Measure')
        Name = Measure.find('Name')
        if Name is None: return {}
        type_ = Measure.attrib['Type']
        if not (type_ == "Deletion" or\
                type_ == "Duplication" or\
                type_ == "copy number gain" or\
                type_ == "copy number loss"):
            return {}
        clinsign = RCVA.find("ClinicalSignificance").find("Description").text
        if clinsign.lower().find("pathogenic") == -1: return {}
        CL = Measure.find("CytogeneticLocation")
        Origin = RCVA.find("ObservedIn").find("Sample").find("Origin")
        TS = RCVA.find("TraitSet").find("Trait")

        allele_id = Measure.attrib['ID']
        date_update = CVA.attrib['DateUpdated']
        origin = Origin.text
        name = Name.find('ElementValue').text
        if CL is not None:
            cytogenetic = CL.text
        else:
            cytogenetic = ""
        seq_locations = Measure.findall("SequenceLocation")
        (assembly, chrsymbol, start, end) = ('', '', -2, -2)
        if not seq_locations:
            seq_locations = list(Measure.iterdescendants("SequenceLocation"))

        if not len(seq_locations) and self.debug:
            print "\n1. %s does not have SequenceLocation\n" % rcv_accession
        for seq_loc in seq_locations:
            if seq_loc.attrib['Assembly'] == 'GRCh37':
                chrsymbol = seq_loc.attrib['Chr']
                assembly = seq_loc.attrib['Assembly']
                start = seq_locations[0].attrib.get("innerStart",seq_locations[0].attrib.get("start",-1))
                end = seq_locations[0].attrib.get("innerStop",seq_locations[0].attrib.get("stop",-1))
                break
        pubmeds = []
        for item in CVAs:
            for cit in item.iterdescendants("Citation"):
                idd = cit.find("ID")
                if idd is not None and idd.attrib['Source'] == 'PubMed':
                    pubmeds.append(idd.text)
        pubs = ",".join(pubmeds)
        genereview = ""
        if TS.find('AttributeSet') is not None and TS.find('AttributeSet').find("XRef") is not None:
            if TS.find('AttributeSet').find("XRef").attrib['DB'] == 'GeneReviews':
                genereview = TS.find('AttributeSet').find("XRef").attrib['ID']
        if (start == -2 or end == -2) and self.debug:
            print "\n2. %s start and end have not been modified" % rcv_accession
        if (start == -1 or end == -1) and self.debug:
            print "\n3. %s does not find start and end in node" % rcv_accession
        doc_hash = {
            "allele_id": allele_id, "type": type_, "name": name,
            "rcv_accession": rcv_accession, "clinsign":clinsign, "origin": origin,
            "assembly": assembly, "chr": chrsymbol, "start": start, 'end': end,
            "cytogenetic":cytogenetic, "date_update":date_update,"pubmeds":pubs,
            "gene_reviews":genereview,
        }
        return doc_hash
    def next(self):
        return self.__next__()
    def __next__(self):
        print "good"

class ParseGeneReview(object):
    """
    下载ftp://ftp.ncbi.nih.gov/pub/GeneReviews/GRshortname_NBKid_genesymbol_dzname.txt 文件，并解析出需要的内容
    """
    def __init__(self, url, path, filename, debug = False, nthread = 10):
        self.url = url
        self.path = path
        self.filename = filename
        self.ftp = FTP(url)
        self._data = []
        self.queue = Queue()
        self.thread = []
        self.http = urllib3.PoolManager()
        self.debug = debug
        self.nthread = nthread
        for i in range(self.nthread):
            worker = Thread(target=self._update_location, args=(i,))
            worker.setDaemon(True)
            self.thread.append(worker)
        if debug:
            print "生成ParserGeneReview对象"

    def _update_location(self, num):
        while True:
            if self.debug: print "Worker %s is working" % num
            if self.queue.qsize() == 0:
                if self.debug: print "队列为空"
                break
            one = self.queue.get()
            url = "http://grch37.rest.ensembl.org/lookup/symbol/homo_sapiens/%s?content-type=application/json" % one['gene_symbol']
            try:
                if self.debug: print "下载%s" % one['gene_symbol']
                tx = self.http.request('GET', url)
                if tx.status == 200:
                    data = json.loads(tx.data.decode('utf-8'))
                    one['chr'] = data['seq_region_name']
                    one['end'] = int(data['end'])
                    one['start'] = int(data['start'])
                elif tx.status == 400:
                    if self.debug: print "%s is not find in web" % (one['gene_symbol'])
                else:
                    raise Exception("下载%s坐标失败" % one['gene_symbol'])
            except Exception,e:
                if self.debug: print "[Error] for get %s\n%s" % (url, e)
            finally:
                self.queue.task_done()

    def _handle_binary(self, more_data):
        self._data.append(more_data)

    def download(self):
        if self.debug: print "Begin downloading"
        self.ftp.login()
        self.ftp.cwd(self.path)
        self.ftp.retrbinary("RETR " + self.filename , callback=self._handle_binary)
        if self.debug:
            print "Finish downloading"
            print "Generate genes dict"
        self.data = "".join(self._data).split("\n")
        if self.data[-1] == "":
            self.data.pop(-1)
        genes = {}
        for item in self.data:
            items = item.split("|")
            if items[2] == '-' or items[2] == 'Not applicable':
                continue
            else:
                try:
                    try:
                        items[3].encode("utf8")
                        genes[items[2]]['accession_0'].append(items[1])
                        genes[items[2]]['description_0'].append(items[3])
                    except UnicodeDecodeError,e:
                        continue
                except KeyError,e:
                    try:
                        items[3].encode("utf8")
                        genes[items[2]] = {"accession_0": [items[1]], 'description_0': [items[3]], 'gene_symbol': items[2]}
                    except UnicodeDecodeError,e:
                        continue
        for key,value in genes.iteritems():
            value['accession'] = ";".join(value['accession_0'])
            value.pop('accession_0')
            value['description'] = ";".join(value['description_0'])
            value.pop("description_0")
        self.genes = genes
        if self.debug:
            print "Finish generation"

    def download_location(self):
        if self.debug:
            print "Begin update location info"
        try:
            self.genes
        except:
            self.download()
        for one in self.genes.itervalues():
            self.queue.put(one)
        for worker in self.thread:
            worker.start()
        for worker in self.thread:
            worker.join()
        self.queue.join()

    def __iter__(self):
        try:
            return self.genes.itervalues()
        except Exception,e:
            self.download()
            self.download_location()
            return self.genes.itervalues()

class ParsePubmed(object):
    """
    通过输入文件解析生成可以导入pubmed表的数据，文件一般有刘二红或李东双给出
    """
    def __init__(self, filename, debug=False):
        if not os.path.exists(filename):
            raise Exception("%s not exists!" % filename)
        self.filename = filename
        self.debug = debug

    def generate_data(self):
        wb = load_workbook(filename = self.filename)
        wh = wb['Sheet1']
        self.data = []
        head = {
                1: 'name', 2: 'gender', 6: 'location', 9: 'gainloss', 10: 'description',
                11: 'pmid', 12: 'cytoband', 13: 'size', 14: 'origin_position', 15: 'critical',
                16: 'hg_ver', 17: 'inheritance', 18: 'have_fulltext', 19: 'note',
                20: 'extra_desc', 21: 'auditor', 22: 'comment'
        }
        for row in wh.iter_rows():
            if row[0].value.find(u"样本信息") != -1 or row[0].value.find(u"编号") != -1:
                continue
            data = {};name = ""
            for cell in row:
                try:
                    head[cell.col_idx]
                except KeyError,e:
                    cellvalue = cell.value if cell.value else 'None'
                    if self.debug: print "Unknown field\n\tIndex -> %s\n\tValue -> %s" % (cell.col_idx, cellvalue)
                    continue
                if cell.col_idx == 1:
                    name = cell.value.encode()
                elif re.search("description|cytoband|size|origin_position|critical|hg_ver|inheritance|note|extra_desc|auditor|comment",head[cell.col_idx]):
                    if cell.value:
                        data[head[cell.col_idx]] = cell.value
                    else:
                        pass
                elif head[cell.col_idx] == 'gender':
                    if cell.value and cell.value.find(u"男") != -1:
                        data['gender'] = 1
                    elif cell.value and cell.value.find(u"女") != -1:
                        data['gender'] = 0
                    elif not cell.value:
                        if self.debug: print "%s lack of gender" % name
                    else:
                        raise Exception("%s name's gender is %s: can not parse" % name, cell.value)
                elif head[cell.col_idx] == 'location':
                    if cell.value:
                        (data['chr'], pos) = cell.value.encode().replace("chr","").split(":")
                        (data['start'], data['end']) = list(map(int, pos.split("-")))
                    else:
                        if self.debug: print "%s lack of 位置" % name
                        data['chr'] = "";data['end'] = 0;data['start'] = 0
                elif head[cell.col_idx] == 'gainloss':
                    if not cell.value:
                        if self.debug: print "%s lack of 突变类型" % name
                    elif cell.value.find(u"重复") != -1:
                        data['gainloss'] = 'gain'
                    elif cell.value.find(u"缺失") != -1:
                        data['gainloss'] = 'loss'
                    else:
                        raise Exception("%s type is %s: can not parser" % name, cell.value)
                elif head[cell.col_idx] == 'pmid':
                    if not cell.value:
                        if self.debug: print "%s lack of PMID号" % name
                    else:
                        data['pmid'] = str(cell.value).split("/")
                elif head[cell.col_idx] == 'have_fulltext':
                    if not cell.value:
                        if self.debug: print "%s lack of 全文判读" % name
                    else:
                        data['have_fulltext'] = True if cell.value.find(u"有") != -1 else False
                else:
                    raise Exception("%s遗漏了一个域%s" % (name, head[cell.col_idx]))
            self.data.append(data)
        return self.data

class ParseCytoband(object):
    def __init__(self, fp, debug=False):
        if not os.path.exists(fp): raise Exception("%s not exists" % fp)
        self.filename = fp
        self.debug = debug
        self.parser_cytoband()
    def parser_cytoband(self):
        if self.debug: print "Parsing filename: %s" % self.filename
        self.cytobands = []
        with open(self.filename) as f:
            for line in f:
                items = line.rstrip().split("\t")
                one = {'chrom': items[0], 'start': int(items[1]), 'end': int(items[2]),
                        'name': items[3], 'description': items[4], 'chr': items[0].replace('chr',"")}
                self.cytobands.append(one)

def myreadlines(f, newline):
  buf = ""
  while True:
    while newline in buf:
      pos = buf.index(newline)
      yield buf[:pos]
      buf = buf[pos + len(newline):]
    chunk = f.read(4096)
    if not chunk:
      yield buf
      break
    buf += chunk

def debugprint(text):
    print "*"*60
    print text;
    print "#"*60


